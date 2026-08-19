#!/usr/bin/env python
"""Regenerate the experiment report workbook from the actual run outputs.

Rewrites Sheet2 (matrix), Sheet3 (per-repeat) and Sheet4 (cross-environment comparison)
of example_files/local_checkpoint_reports_v2.xlsx. Sheet1 is left untouched -- it is the
original hand-written draft and is kept as a record.

Metrics are READ FROM DISK, never typed in: for every tag with a `results_glob` below,
each repeat's metrics_summary.tsv and evaluation.tsv are parsed, and the mean/sd are
computed across repeats. A tag whose runs do not exist yet simply comes out blank, so
this can be re-run after each experiment finishes and the sheet fills itself in.

    python muat/pkg_reproduce/make_report_workbook.py
    python muat/pkg_reproduce/make_report_workbook.py --check   # report, write nothing
"""

import argparse
import csv
import datetime
import glob as globmod
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
WORKBOOK = os.path.join(REPO, 'example_files', 'local_checkpoint_reports_v2.xlsx')

HDR_FILL = PatternFill('solid', fgColor='1F3864')
SUB_FILL = PatternFill('solid', fgColor='2E5496')
DONE_FILL = PatternFill('solid', fgColor='E2EFDA')
TODO_FILL = PatternFill('solid', fgColor='FFF2CC')
BLOCKED_FILL = PatternFill('solid', fgColor='FBE5D6')
SUMMARY_FILL = PatternFill('solid', fgColor='F2F2F2')
HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
NOTE_FONT = Font(size=9, italic=True)
BOLD = Font(bold=True, size=10)
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ACC_FMT = '0.000000'

# ---------------------------------------------------------------------------
# Tag definitions. `results_glob` is what makes a row self-filling; tags without
# one are not runnable yet and stay blank by design.
# ---------------------------------------------------------------------------

COHORT_HG19 = 'PCAWG open-access (hg19)\n1812 samples / 17 classes'
COHORT_HG38 = 'PCAWG open-access\n(hg19 calls lifted to hg38)'
SPLIT_TRAIN = '80:20 - 1449 train / 363 test'
SPLIT_TEST = 'Test (363)'
INPUT_TYPE = 'SNV+MNV+pos+ges'

HP = {'n_emb': 128, 'n_layer': 1, 'n_head': 1, 'lr': 0.001, 'epoch': 100}

TAGS = [
    {
        'tag': 'd0',
        'purpose': 'Reference run - establishes the reported figures and run-to-run determinism',
        'cohort': COHORT_HG19, 'mode': 'Train', 'seed': 1337, 'repeats': 10,
        'split': SPLIT_TRAIN, 'input_type': INPUT_TYPE,
        'ckpt_in': 'None', 'ckpt_out': 'ckpt1 (d1.pthx) + ckpt1.zip',
        'spe': 'internal GPU node (pre-Puhti verification run, not the reported Puhti/bioconda environment)',
        'hardware': 'Tesla P100-PCIE-16GB, driver 530.30.02',
        'install': 'source (repo, muat-gpu-2 env)', 'muat_version': '0.1.20 (pre-release source)',
        'torch': 'torch 2.5.1.post303 / CUDA 12.1',
        'results_glob': 'data/reproduce_results/d1_rep*_280085',
        'ckpt_sha256': '763b6e768752c4acb5d8ea8b9139ad31de282c1bde084f8fe9b902b5a1e38cea',
        'determinism': '10/10 identical - weight tensors, all configs, logits, predictions '
                       'and the full 100-epoch curve (sd 0.000000)',
        'status': 'DONE 2026-08-07 (SLURM array 280085)',
        'state': 'done',
    },
    {
        'tag': 'd1',
        'purpose': 'Training-run reproducibility - same environment',
        'cohort': COHORT_HG19, 'mode': 'Train', 'seed': 1337, 'repeats': 10,
        'split': SPLIT_TRAIN, 'input_type': INPUT_TYPE,
        'ckpt_in': 'None', 'ckpt_out': 'ckpt1 (d1.pthx) + ckpt1.zip -> Zenodo',
        'spe': 'CSC Puhti GPU',
        'hardware': 'Tesla P100-PCIE-16GB, driver 530.30.02',
        'install': 'installed from bioconda (muat 0.1.22), no pip step',
        'muat_version': '0.1.22', 'torch': '',
        'results_glob': None,
        'determinism': '',
        'status': 'TO RUN - run_reproduce_d1_bioconda_x10.sbatch (array 1-10%1)',
        'state': 'todo',
    },
    {
        'tag': 'd2',
        'purpose': 'Training-run reproducibility - different environment',
        'cohort': COHORT_HG19, 'mode': 'Train', 'seed': 1337, 'repeats': 10,
        'split': SPLIT_TRAIN, 'input_type': INPUT_TYPE,
        'ckpt_in': 'None', 'ckpt_out': 'ckpt2 (d2.pthx) + ckpt2.zip',
        'spe': 'CSC Puhti CPU',
        'hardware': 'CPU only, 4 threads/repeat, MKL_CBWR=AVX2',
        'install': 'published BioContainers image muat:0.1.22--pyh106432d_0 (built by bioconda from the v0.1.22 release tarball), executed by apptainer',
        'muat_version': '0.1.22', 'torch': '',
        'results_glob': None,
        'determinism': '',
        'status': 'TO RUN - run_reproduce_d2_x10.sbatch (array 1-10), concurrent with d1',
        'state': 'todo',
    },
    {
        'tag': 'd3',
        'purpose': 'Inference-run reproducibility - same environment',
        'cohort': COHORT_HG19, 'mode': 'Inference', 'seed': 1337, 'repeats': 10,
        'split': SPLIT_TEST, 'input_type': INPUT_TYPE,
        'ckpt_in': 'ckpt1', 'ckpt_out': 'None',
        'spe': 'CSC Puhti GPU',
        'hardware': 'Tesla P100-PCIE-16GB, driver 530.30.02',
        'install': 'installed from bioconda (muat 0.1.22), no pip step',
        'muat_version': '0.1.22', 'torch': '',
        'results_glob': None,
        'determinism': '',
        'status': 'TO RUN - run_reproduce_d3.sbatch (single job, repeat x10 internal); needs ckpt1',
        'state': 'todo',
    },
    {
        'tag': 'd4',
        'purpose': 'Inference-run reproducibility - different environment',
        'cohort': COHORT_HG19, 'mode': 'Inference', 'seed': 1337, 'repeats': 10,
        'split': SPLIT_TEST, 'input_type': INPUT_TYPE,
        'ckpt_in': 'ckpt1', 'ckpt_out': 'None',
        'spe': 'CSC Puhti CPU',
        'hardware': 'CPU only, 4 threads/repeat, MKL_CBWR=AVX2',
        'install': 'published BioContainers image muat:0.1.22--pyh106432d_0 (built by bioconda from the v0.1.22 release tarball), executed by apptainer',
        'muat_version': '0.1.22', 'torch': '',
        'results_glob': None,
        'determinism': '',
        'status': 'TO RUN - run_reproduce_d4.sbatch (single job, repeat x10 internal); needs ckpt1',
        'state': 'todo',
    },
    {
        'tag': 'd5',
        'purpose': 'Training with GRCh38 (hg38)',
        'cohort': COHORT_HG38, 'mode': 'Train', 'seed': 1337, 'repeats': 10,
        'split': SPLIT_TRAIN, 'input_type': INPUT_TYPE,
        'ckpt_in': 'None', 'ckpt_out': 'ckpt3 (d5.pthx) + ckpt3.zip',
        'spe': 'CSC Puhti GPU', 'hardware': '', 'install': 'bioconda',
        'muat_version': '0.1.22', 'torch': '',
        'results_glob': None,
        'determinism': '', 'status': 'BLOCKED - see note 8', 'state': 'blocked',
    },
    {
        'tag': 'd6',
        'purpose': 'Inference with GRCh38 (hg38)',
        'cohort': COHORT_HG38, 'mode': 'Inference', 'seed': 1337, 'repeats': 10,
        'split': SPLIT_TEST, 'input_type': INPUT_TYPE,
        'ckpt_in': 'ckpt3', 'ckpt_out': 'None',
        'spe': 'CSC Puhti GPU', 'hardware': '', 'install': 'bioconda',
        'muat_version': '0.1.22', 'torch': '',
        'results_glob': None,
        'determinism': '', 'status': 'BLOCKED - depends on d5', 'state': 'blocked',
    },
]

METRIC_KEYS = [
    ('top1_accuracy', 'Top-1 acc'),
    ('top3_accuracy', 'Top-3 acc'),
    ('top5_accuracy', 'Top-5 acc'),
    ('weighted_precision', 'precision (weighted)'),
    ('weighted_recall', 'recall (weighted)'),
    ('weighted_f1', 'weighted-F1'),
    ('macro_f1', 'macro-F1'),
]

NOTES = [
    'How to read this sheet:',
    '1. This sheet is GENERATED by muat/pkg_reproduce/make_report_workbook.py, which reads the '
    'metrics straight out of each run directory. Do not type numbers in by hand - re-run the '
    'script after an experiment finishes and the row fills itself. Sheet1 is the original '
    'hand-written draft, kept unchanged for reference.',
    '2. Values are repeated on every row rather than merged across rows. Merged cells were '
    'previously misread as missing data; explicit repetition removes the ambiguity.',
    '3. Every tag uses seed 1337 deliberately. Attributing a difference to the ENVIRONMENT '
    'requires the seed to be held fixed; varying both confounds seed and environment.',
    '4. WITHIN an environment the repeats of one tag are expected IDENTICAL. ACROSS environments '
    '(d1 vs d2, d3 vs d4) they will NOT be bit-identical - CPU and GPU use different kernels and '
    'reduction orders. The agreed tolerance is defined on Sheet4 and was fixed BEFORE running.',
    '5. d0 vs d1 is the install axis (source vs bioconda) on the same hardware; d1 vs d2 changes '
    'device (GPU->CPU) and install (bioconda->docker) together. If a d1/d2 discrepancy needs '
    'attributing, add an intermediate Puhti GPU + docker run to separate container from device.',
    '6. d1 and d2 use the SAME muat version (0.1.22): one installed via bioconda, one from the '
    'PUBLISHED BioContainers image for that same release - not an image we built, so the '
    'artefact is one a reader can pull. Both arms assert the version at runtime; if the '
    'versions differ the comparison is void.',
    '7. d3/d4/d6 evaluate on d1\'s own 363-sample test split. This is an environment fingerprint, '
    'NOT a held-out generalisation estimate, and must not be presented as one.',
    '8. Best epoch is selected on those same 363 samples (trainer.py:324), so the test set '
    'informed epoch selection. Report these as best-epoch figures. "Best epoch" is N/A for the '
    'inference tags, which run a fixed checkpoint.',
    '9. d5/d6 are not runnable yet: no hg38 bundle exists, and the shipped position dictionary '
    '(dictChpos.tsv) is hg19-derived - see the _hg38_note on tag d5 in experiments.json.',
    '10. CAVEAT for d5/d6: PCAWG open-access variants were called against hg19, so the hg38 arm '
    'runs on lifted-over coordinates. It therefore measures the effect of a liftover round-trip, '
    'not performance on natively hg38-called data. State this rather than let a reviewer say it.',
]

SHEET4_TOLERANCE = [
    ('Tolerance, fixed before any cross-environment run was executed', ''),
    ('Prediction agreement', 'at least 362/363 top-1 predictions identical (delta Top-1 acc <= 0.003)'),
    ('Logit deviation', 'max |delta logit| < 1e-3 over all 363 x 17 logits'),
    ('Disagreements', 'any flipped sample must be a near-tie: its top-2 logit margin in the '
                      'reference run must be smaller than the logit perturbation observed on '
                      'that same sample. Judged per sample, not against the global maximum, so '
                      'one outlier cannot excuse every flip.'),
    ('Rationale', 'CPU and GPU float32 kernels differ in reduction order, so bit-equality across '
                  'devices is not achievable and should not be claimed. A near-tie flip is '
                  'numerical noise; a confident-prediction flip is a real portability defect.'),
]

SHEET4_PAIRS = [
    ('d0 vs d1', 'install: source -> bioconda (same GPU hardware)'),
    ('d1 vs d2', 'device GPU -> CPU AND install bioconda -> docker'),
    ('d3 vs d4', 'device GPU -> CPU AND install bioconda -> docker (inference, same ckpt1)'),
    ('d1 vs d5', 'genome build hg19 -> hg38 (NOT a portability test - a different experiment)'),
]


# ---------------------------------------------------------------------------
# Reading run outputs
# ---------------------------------------------------------------------------

def read_metrics_summary(path):
    out = {}
    with open(path) as fh:
        for row in csv.DictReader(fh, delimiter='\t'):
            try:
                out[row['metric']] = float(row['value'])
            except (KeyError, TypeError, ValueError):
                continue
    return out


def read_best_epoch(path):
    """Epoch with the highest validation_accuracy in evaluation.tsv (1-indexed)."""
    if not os.path.isfile(path):
        return None
    best, best_idx = None, None
    with open(path) as fh:
        for idx, row in enumerate(csv.DictReader(fh, delimiter='\t'), start=1):
            key = next((k for k in row if k and 'val' in k.lower() and 'acc' in k.lower()), None)
            if key is None:
                return None
            try:
                val = float(row[key])
            except (TypeError, ValueError):
                continue
            if best is None or val > best:
                best, best_idx = val, idx
    return best_idx


# ---------------------------------------------------------------------------
# Compute cost (answers reviewer 1 minor 8)
#
# Read from the job logs rather than typed into the tag table, for the same reason the metrics
# are: a hand-entered runtime silently goes stale. Wall-clock is derived from the
# "=== ... starting <date> ===" / "finished" markers instead of the "wall clock :" printf,
# because the markers exist in EVERY log -- including d0's array 280085 and the two arrays
# submitted before the cost instrumentation was added, whose scripts sbatch had already
# snapshotted.
# ---------------------------------------------------------------------------

RESULTS_DIR = os.path.join(REPO, 'data', 'reproduce_results')

_RE_RESULT_DIR = re.compile(r'^result dir\s*:\s*(\S+)', re.M)
_RE_STAMP = re.compile(r'^=== reproduce \S+(?: repeat \d+)? (starting|finished) (.+?) ===', re.M)
_RE_PEAK_GPU = re.compile(r'^peak GPU mem:\s*(\d+)\s*MiB', re.M)
# sacct --units=M rows, e.g. "  280271_1.batch  d1_bioco  01:29:12  05:12.3  1234.56M  ...".
# The JobID column is anchored to a real SLURM id, and only the text after the "slurm cost :"
# marker is searched. Both guards are needed: a looser pattern matched the `ls -lh` line at the
# end of each log ("... 5.4M Aug  6 ...") and silently reported a 0.0 GB peak.
_RE_SACCT_MARKER = re.compile(r'^slurm cost\s*:', re.M)
_RE_SACCT_RSS = re.compile(
    r'^\s*\d+(?:_\d+)?(?:\.\S+)?\s+\S+\s+\S+\s+\S+\s+([\d.]+)M(?=\s|$)', re.M)


def _parse_stamp(text):
    """`Thu Aug  6 11:25:55 EEST 2026` -> datetime. The timezone NAME is dropped: %Z only
    accepts the running process's own zone, so parsing it fails on a machine set elsewhere."""
    parts = text.split()
    if len(parts) == 6:
        parts.pop(4)
    try:
        return datetime.datetime.strptime(' '.join(parts), '%a %b %d %H:%M:%S %Y')
    except ValueError:
        return None


def read_costs():
    """Map result-dir (repo-relative) -> {'wall_s', 'gpu_mib', 'rss_mb'}, from the job logs.

    Every log names its own result dir, so the mapping needs no knowledge of job-name
    conventions or array ids.
    """
    costs = {}
    for log in sorted(globmod.glob(os.path.join(RESULTS_DIR, '*.out'))):
        try:
            with open(log, errors='replace') as fh:
                text = fh.read()
        except OSError:
            continue
        m = _RE_RESULT_DIR.search(text)
        if not m:
            continue
        key = os.path.relpath(m.group(1), REPO)
        rec = costs.setdefault(key, {'wall_s': None, 'gpu_mib': None, 'rss_mb': None})

        stamps = {}
        for kind, raw in _RE_STAMP.findall(text):
            stamp = _parse_stamp(raw)
            if stamp:
                stamps[kind] = stamp
        if 'starting' in stamps and 'finished' in stamps:
            delta = (stamps['finished'] - stamps['starting']).total_seconds()
            if delta >= 0:
                rec['wall_s'] = delta

        gpu = _RE_PEAK_GPU.search(text)
        if gpu:
            rec['gpu_mib'] = int(gpu.group(1))
        # Largest MaxRSS across the sacct rows (the .batch step carries the real value),
        # searched only inside the "slurm cost :" section.
        marker = _RE_SACCT_MARKER.search(text)
        if marker:
            rss = [float(v) for v in _RE_SACCT_RSS.findall(text[marker.end():])]
            if rss:
                rec['rss_mb'] = max(rss)
    return costs


def fmt_wall(seconds):
    if not seconds:
        return ''
    seconds = int(round(seconds))
    return '%d:%02d:%02d' % (seconds // 3600, seconds % 3600 // 60, seconds % 60)


def summarise_cost(repeats):
    """('wall-clock per repeat', 'peak memory') for one tag. Blank when unmeasured."""
    walls = [r['_cost']['wall_s'] for r in repeats if r.get('_cost', {}).get('wall_s')]
    gpus = [r['_cost']['gpu_mib'] for r in repeats if r.get('_cost', {}).get('gpu_mib')]
    rss = [r['_cost']['rss_mb'] for r in repeats if r.get('_cost', {}).get('rss_mb')]

    wall = ''
    if walls:
        wall = fmt_wall(sum(walls) / len(walls))
        if len(walls) > 1:
            # Spread matters: a wide range means the node was contended, which is worth
            # seeing next to a determinism claim.
            wall += ' (mean of %d; %s-%s)' % (len(walls), fmt_wall(min(walls)), fmt_wall(max(walls)))
    mem = []
    if gpus:
        mem.append('%.1f GB GPU' % (max(gpus) / 1024.0))
    if rss:
        mem.append('%.1f GB RSS' % (max(rss) / 1024.0))
    return wall, ' / '.join(mem)


def collect(results_glob, costs=None):
    """Per-repeat metric dicts for one tag, in repeat order. Empty if nothing ran."""
    if not results_glob:
        return []
    costs = costs or {}
    repeats = []
    for d in sorted(x for x in globmod.glob(os.path.join(REPO, results_glob)) if os.path.isdir(x)):
        summary = os.path.join(d, 'metrics_summary.tsv')
        if not os.path.isfile(summary):
            continue
        rec = read_metrics_summary(summary)
        rec['_dir'] = os.path.relpath(d, REPO)
        rec['_best_epoch'] = read_best_epoch(os.path.join(d, 'evaluation.tsv'))
        rec['_cost'] = costs.get(rec['_dir'], {'wall_s': None, 'gpu_mib': None, 'rss_mb': None})
        repeats.append(rec)
    return repeats


def mean_sd(values):
    """Mean and POPULATION sd -- these are repeats of one fixed configuration, not a sample."""
    vals = [v for v in values if v is not None]
    if not vals:
        return None, None
    m = sum(vals) / len(vals)
    return m, (sum((v - m) ** 2 for v in vals) / len(vals)) ** 0.5


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def style_header(ws, row, values, fill=HDR_FILL):
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill, c.font, c.border = fill, HDR_FONT, BOX
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def write_sheet2(wb, data):
    if 'Sheet2' in wb.sheetnames:
        del wb['Sheet2']
    ws = wb.create_sheet('Sheet2', 0)

    groups = [
        ('', ['Tag', 'Purpose', 'Cohort', 'Mode', 'Seed', 'Repeats']),
        ('Hyperparameters', ['n_emb', 'n_layer', 'n_head', 'learning rate', 'epoch']),
        ('', ['Split', 'Input Type', 'Checkpoint in', 'Checkpoint produced']),
        ('Environment', ['SPE', 'Hardware', 'Installation', 'muat version', 'torch / CUDA']),
        ('Artifact', ['Checkpoint sha256']),
        ('Cost', ['Wall-clock / repeat', 'Peak memory']),
        ('Performance metrics (mean over repeats, best epoch)',
         [label for _, label in METRIC_KEYS] + ['Best epoch']),
        ('', ['Determinism (within tag)', 'Status']),
    ]

    col = 1
    spans = {}
    for group, headers in groups:
        for h in headers:
            spans[h] = col
            if group:
                ws.cell(row=2, column=col, value=h)
            else:
                ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
                ws.cell(row=1, column=col, value=h)
            col += 1
        if group:
            start = col - len(headers)
            if len(headers) > 1:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col - 1)
            ws.cell(row=1, column=start, value=group)

    for r in (1, 2):
        for c in range(1, col):
            cell = ws.cell(row=r, column=c)
            cell.fill = HDR_FILL if r == 1 else SUB_FILL
            cell.font, cell.border = HDR_FONT, BOX
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    fills = {'done': DONE_FILL, 'todo': TODO_FILL, 'blocked': BLOCKED_FILL}
    row = 3
    for spec in TAGS:
        agg = data[spec['tag']]
        measured_wall, measured_mem = summarise_cost(agg)
        vals = {
            'Tag': spec['tag'], 'Purpose': spec['purpose'], 'Cohort': spec['cohort'],
            'Mode': spec['mode'], 'Seed': spec['seed'], 'Repeats': spec['repeats'],
            'n_emb': HP['n_emb'], 'n_layer': HP['n_layer'], 'n_head': HP['n_head'],
            'learning rate': HP['lr'], 'epoch': HP['epoch'],
            'Split': spec['split'], 'Input Type': spec['input_type'],
            'Checkpoint in': spec['ckpt_in'], 'Checkpoint produced': spec['ckpt_out'],
            'SPE': spec['spe'], 'Hardware': spec['hardware'], 'Installation': spec['install'],
            'muat version': spec['muat_version'], 'torch / CUDA': spec['torch'],
            'Checkpoint sha256': spec.get('ckpt_sha256', ''),
            # Measured from the logs where possible; the tag table is only a fallback for
            # rows that have not run yet.
            'Wall-clock / repeat': measured_wall or spec.get('wall_clock', ''),
            'Peak memory': measured_mem or spec.get('peak_memory', ''),
            'Determinism (within tag)': spec['determinism'], 'Status': spec['status'],
        }
        for key, label in METRIC_KEYS:
            m, _ = mean_sd([r.get(key) for r in agg])
            vals[label] = m
        if spec['mode'] == 'Inference':
            vals['Best epoch'] = 'N/A (fixed checkpoint)'
        else:
            be, _ = mean_sd([r.get('_best_epoch') for r in agg])
            vals['Best epoch'] = int(be) if be is not None else None

        for header, c in spans.items():
            cell = ws.cell(row=row, column=c, value=vals.get(header))
            cell.border, cell.fill = BOX, fills[spec['state']]
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if header in [l for _, l in METRIC_KEYS]:
                cell.number_format = ACC_FMT
        ws.cell(row=row, column=spans['Tag']).font = BOLD
        row += 1

    row += 1
    for i, note in enumerate(NOTES):
        cell = ws.cell(row=row, column=1, value=note)
        cell.font = BOLD if i == 0 else NOTE_FONT
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        ws.row_dimensions[row].height = 28 if i else 15
        row += 1

    widths = {'Tag': 6, 'Purpose': 40, 'Cohort': 26, 'Mode': 10, 'Seed': 7, 'Repeats': 8,
              'Split': 24, 'Input Type': 17, 'Checkpoint in': 13, 'Checkpoint produced': 24,
              'SPE': 18, 'Hardware': 34, 'Installation': 24, 'muat version': 20,
              'torch / CUDA': 20, 'Checkpoint sha256': 22, 'Wall-clock / repeat': 16,
              'Peak memory': 13, 'Best epoch': 13,
              'Determinism (within tag)': 42, 'Status': 34}
    for header, c in spans.items():
        ws.column_dimensions[get_column_letter(c)].width = widths.get(header, 12)
    ws.freeze_panes = 'B3'
    return ws


def write_sheet3(wb, data):
    if 'Sheet3' in wb.sheetnames:
        del wb['Sheet3']
    ws = wb.create_sheet('Sheet3', 1)

    headers = ['Tag', 'Run', 'Result directory'] + [l for _, l in METRIC_KEYS] + ['Best epoch']
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value='Tag')
    ws.cell(row=1, column=2, value='Run')
    ws.cell(row=1, column=3, value='Result directory')
    ws.cell(row=1, column=4, value='Performance metrics (best epoch)')
    for c, h in enumerate(headers[3:], start=4):
        ws.cell(row=2, column=c, value=h)
    for r in (1, 2):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = HDR_FILL if r == 1 else SUB_FILL
            cell.font, cell.border = HDR_FONT, BOX
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    row = 3
    for spec in TAGS:
        tag, n = spec['tag'], spec['repeats']
        agg = data[tag]
        start = row
        for i in range(n):
            rec = agg[i] if i < len(agg) else {}
            ws.cell(row=row, column=2, value=i + 1)
            ws.cell(row=row, column=3, value=rec.get('_dir', ''))
            for j, (key, _) in enumerate(METRIC_KEYS, start=4):
                cell = ws.cell(row=row, column=j, value=rec.get(key))
                cell.number_format = ACC_FMT
            if spec['mode'] == 'Inference':
                ws.cell(row=row, column=4 + len(METRIC_KEYS), value='N/A' if rec else None)
            else:
                ws.cell(row=row, column=4 + len(METRIC_KEYS), value=rec.get('_best_epoch'))
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = BOX
            row += 1
        ws.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
        tcell = ws.cell(row=start, column=1, value=tag)
        tcell.font = BOLD
        tcell.alignment = Alignment(horizontal='center', vertical='center')

        for stat in ('mean', 'sd'):
            ws.cell(row=row, column=2, value=stat)
            ws.cell(row=row, column=3, value='%d repeat(s) found' % len(agg))
            for j, (key, _) in enumerate(METRIC_KEYS, start=4):
                m, s = mean_sd([r.get(key) for r in agg])
                cell = ws.cell(row=row, column=j, value=m if stat == 'mean' else s)
                cell.number_format = ACC_FMT
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill, cell.border, cell.font = SUMMARY_FILL, BOX, BOLD
            row += 1
        row += 1

    ws.cell(row=row, column=1,
            value='sd is the POPULATION standard deviation over repeats of one fixed '
                  'configuration. sd = 0.000000 means the repeats are identical, not that the '
                  'rows were copy-pasted.').font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

    for c, w in enumerate([6, 8, 34] + [13] * len(METRIC_KEYS) + [11], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'C3'
    return ws


def write_sheet4(wb):
    if 'Sheet4' in wb.sheetnames:
        del wb['Sheet4']
    ws = wb.create_sheet('Sheet4', 2)

    ws.cell(row=1, column=1, value='Cross-environment comparison').font = Font(bold=True, size=12)
    row = 3
    for label, text in SHEET4_TOLERANCE:
        ws.cell(row=row, column=1, value=label).font = BOLD
        cell = ws.cell(row=row, column=2, value=text)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 30 if text else 16
        row += 1

    row += 1
    headers = ['Pair', 'What differs', 'Predictions agreeing (/363)', 'Agreement (%)',
               'delta Top-1 acc', 'max |delta logit|', 'mean |delta logit|',
               'All flips near-ties?', 'Tolerance met?', 'Notes']
    style_header(ws, row, headers)
    row += 1
    for pair, differs in SHEET4_PAIRS:
        ws.cell(row=row, column=1, value=pair).font = BOLD
        ws.cell(row=row, column=2, value=differs).alignment = Alignment(vertical='top', wrap_text=True)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            cell.border, cell.fill = BOX, TODO_FILL
        row += 1

    row += 1
    for note in [
        'Populate this sheet with muat/pkg_reproduce/compare_environments.py once both arms of '
        'a pair have run. Bit-equality is NOT the criterion across devices - the tolerance above is.',
        'd1 vs d5 is listed for completeness but is a genome-build experiment, not a portability '
        'test; the tolerance above does not apply to it and a real accuracy difference is expected.',
    ]:
        cell = ws.cell(row=row, column=1, value=note)
        cell.font = NOTE_FONT
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ws.row_dimensions[row].height = 28
        row += 1

    for c, w in enumerate([12, 46, 16, 13, 14, 15, 15, 14, 14, 30], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--workbook', default=WORKBOOK)
    ap.add_argument('--check', action='store_true', help='report what was found, write nothing')
    args = ap.parse_args()

    costs = read_costs()
    data = {}
    for spec in TAGS:
        agg = collect(spec['results_glob'], costs)
        data[spec['tag']] = agg
        found = '%d repeat(s)' % len(agg) if agg else 'no runs yet'
        note = ''
        if agg:
            m, s = mean_sd([r.get('top1_accuracy') for r in agg])
            note = '  top1 mean %.6f sd %.6f' % (m, s)
            wall, mem = summarise_cost(agg)
            if wall or mem:
                note += '  | %s %s' % (wall, mem)
        print('%-4s %-14s %s' % (spec['tag'], found, note))
        if agg and len(agg) != spec['repeats']:
            print('     WARNING: expected %d repeats, found %d' % (spec['repeats'], len(agg)))

    if args.check:
        print('\n--check: nothing written')
        return 0

    # The workbook is a generated artifact and is NOT tracked in git, so a fresh clone
    # will not have one -- build it from scratch in that case. When it does exist it is
    # loaded rather than replaced, to preserve Sheet1 (the original hand-written draft).
    if os.path.isfile(args.workbook):
        wb = openpyxl.load_workbook(args.workbook)
    else:
        print('no workbook at %s -- creating one' % args.workbook)
        wb = openpyxl.Workbook()
        for name in list(wb.sheetnames):          # drop openpyxl's default empty sheet
            del wb[name]
        os.makedirs(os.path.dirname(os.path.abspath(args.workbook)), exist_ok=True)

    write_sheet2(wb, data)
    write_sheet3(wb, data)
    write_sheet4(wb)
    # Keep Sheet1 (the original hand-written draft) last and untouched.
    if 'Sheet1' in wb.sheetnames:
        wb.move_sheet('Sheet1', offset=len(wb.sheetnames))
    wb.active = 0
    wb.save(args.workbook)
    print('\nwrote %s' % args.workbook)
    print('sheets: %s' % ', '.join(wb.sheetnames))
    return 0


if __name__ == '__main__':
    sys.exit(main())
